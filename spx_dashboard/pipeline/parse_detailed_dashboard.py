"""
Parse the Summary tab of the team's Detailed Dashboard workbook into
data/equities_seed.json — the one-time seed (and re-seedable backup) for the
Equities Dashboard page.

The workbook mixes three kinds of cells:
  * analyst model inputs (revs, GM%, EPS, DPS, multiples, share counts...)
    → extracted as values and become the editable model in Supabase;
  * Bloomberg formulas (price, 1M/3M/6M performance, ADV, index BEst P/E)
    → cached values are kept where present (prices are NOT cached when the
      file was saved off-terminal; the site fetches live quotes instead);
  * derived columns (EV/GP, Mendo P/E, IRR, MoM, IRR decomp, CAGRs)
    → NOT extracted; the site recomputes them from the inputs, so they stay
      consistent as prices move and analysts edit the model.

Per-row formula variants (how target price is built, how div yield is
defined, whether the IRR decomp is the full version or the simplified one
used for unprofitable names) are detected from the formula text and stored
as flags so lib/equities/calc.ts can reproduce each row faithfully.

Usage:
    python parse_detailed_dashboard.py <Detailed_Dashboard.xlsx> <equities_seed.json>
"""

from __future__ import annotations

import datetime as dt
import json
import sys

import openpyxl

# Column → year layouts (fixed template, addressed by letter like parse_excel.py).
YEARS_9 = list(range(2022, 2031))  # nine-year series
YEARS_5 = list(range(2026, 2031))  # five-year series

SERIES = {
    "revs": (["O", "P", "Q", "R", "S", "T", "U", "V", "W"], YEARS_9),
    "gm": (["X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"], YEARS_9),
    "gp": (["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO"], YEARS_9),
    "adj_eps": (["AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX"], YEARS_9),
    "mendo_eps": (["AY", "AZ", "BA", "BB", "BC", "BD", "BE", "BF", "BG"], YEARS_9),
    "target_mult": (["CD", "CE", "CF", "CG", "CH"], YEARS_5),
    "ncps": (["CS", "CT", "CU", "CV", "CW"], YEARS_5),
    "wadso": (["DC", "DD", "DE", "DF", "DG"], YEARS_5),
    "net_debt": (["DH", "DI", "DJ", "DK", "DL"], YEARS_5),
    "dps": (["DO", "DP", "DQ", "DR", "DS"], YEARS_5),
}

# Bloomberg ticker suffix → (yahoo suffix, currency symbol, price scale).
# LSE quotes arrive from Yahoo in pence while the model is in pounds.
EXCHANGES = {
    "US": ("", "$", 1.0),
    "GY": (".DE", "€", 1.0),
    "FP": (".PA", "€", 1.0),
    "LN": (".L", "£", 0.01),
    "DC": (".CO", "", 1.0),
    "TT": (".TW", "$", 1.0),
}

INDEX_YAHOO = {"SPX": "^GSPC"}  # B500XM7T is a custom Bloomberg index — no public feed


def num(v):
    """Cached cell → float, or None for blanks / 'n/a' / Bloomberg errors."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None  # strings: 'n/a', '#N/A Connection', '#VALUE!', ...


def series(ws, row, key):
    cols, years = SERIES[key]
    out = {}
    for col, yr in zip(cols, years):
        v = num(ws[f"{col}{row}"].value)
        if v is not None:
            out[str(yr)] = v
    return out


def formula_text(ws_raw, ref):
    v = ws_raw[ref].value
    if hasattr(v, "text"):  # ArrayFormula
        v = v.text
    return v if isinstance(v, str) else ""


def main(src: str, dest: str) -> None:
    raw = openpyxl.load_workbook(src, data_only=False)["Summary"]
    ws = openpyxl.load_workbook(src, data_only=True)["Summary"]

    groups: list[dict] = []
    indexes: list[dict] = []
    current = None

    for r in range(6, ws.max_row + 1):
        bbg = ws[f"B{r}"].value
        label = ws[f"D{r}"].value
        if bbg is None:
            if isinstance(label, str) and label.strip():
                current = {"name": label.strip(), "companies": []}
                if label.strip().lower() != "index":
                    groups.append(current)
                else:
                    current = "INDEX"
            continue

        bbg = str(bbg).strip()
        ticker = str(label).strip()

        if current == "INDEX":
            best_pe = {}
            for col, yr in zip(["BY", "BZ", "CA"], [2026, 2027, 2028]):
                v = num(ws[f"{col}{r}"].value)
                if v is not None:
                    best_pe[str(yr)] = round(v, 4)
            indexes.append(
                {
                    "ticker": ticker,
                    "bbg": bbg,
                    "yahoo": INDEX_YAHOO.get(ticker),
                    "best_pe": best_pe,
                    "perf": {
                        "m1": num(ws[f"EB{r}"].value),
                        "m3": num(ws[f"EC{r}"].value),
                        "m6": num(ws[f"ED{r}"].value),
                    },
                }
            )
            continue

        suffix = bbg.upper().replace(" EQUITY", "").split()[-1]
        y_suffix, ccy, scale = EXCHANGES.get(suffix, ("", "$", 1.0))
        symbol = ticker.split()[0] + y_suffix
        if suffix == "TT":
            symbol = bbg.split()[0] + y_suffix  # numeric Taiwan codes (2330.TW)

        model = {k: series(ws, r, k) for k in SERIES}
        # GM% is canonical; where the sheet hardcodes GP and derives GM, back-fill
        # GM = GP / Revs so the site can always recompute GP = GM * Revs.
        for yr, gp in model["gp"].items():
            if yr not in model["gm"] and model["revs"].get(yr):
                model["gm"][yr] = gp / model["revs"][yr]
        del model["gp"]

        model["shares"] = num(ws[f"I{r}"].value)
        model["cash"] = num(ws[f"J{r}"].value)
        model["debt"] = num(ws[f"K{r}"].value)
        model["min_int"] = num(ws[f"L{r}"].value)

        # Row-specific formula variants -------------------------------------
        cn = formula_text(raw, f"CN{r}")
        if "(CX" in cn:
            variant = "gp_ev"  # (mult·GP − net debt) / WADSO
        elif "AL" in cn and "/DD" in cn:
            variant = "gp_ps"  # mult·GP / WADSO + net cash ps
        elif "(T" in cn:
            variant = "rev_ps"  # mult·Revs / WADSO + net cash ps
        else:
            variant = "pe"  # mult·Mendo EPS (+ net cash ps)
        cash_in_target = "+CS" in cn.replace(" ", "")

        eq = formula_text(raw, f"EQ{r}")
        div_yield_mode = "dps" if "AVERAGE" in eq else ("cashbuild" if "CU" in eq else "none")

        # The decomp wiring varies by row:
        #   standard   – EPS+Divs = mEPS CAGR + div yield; Multiple is residual
        #   mult_first – Multiple = P/E→target-mult CAGR; EPS+Divs is residual
        #   simple     – Yield is a hardcoded assumption; EPS+Divs = Revs + Yield
        #   none       – decomp left blank in the sheet
        ei = formula_text(raw, f"EI{r}")
        if "=ER" in ei:
            decomp = "standard"
        elif "=EK" in ei:
            decomp = "mult_first"
        elif "=EH" in ei:
            decomp = "simple"
        else:
            decomp = "none"
        yield_input = num(ws[f"EH{r}"].value) if decomp == "simple" else None

        upd = ws[f"F{r}"].value
        company = {
            "ticker": ticker,
            "bbg": bbg,
            "yahoo": symbol,
            "currency": ccy,
            "px_scale": scale,
            "port": ws[f"E{r}"].value if isinstance(ws[f"E{r}"].value, (int, float)) else None,
            "update_date": upd.date().isoformat() if isinstance(upd, dt.datetime) else None,
            "update_by": str(ws[f"G{r}"].value).strip() if ws[f"G{r}"].value else None,
            "variant": variant,
            "cash_in_target": cash_in_target,
            "div_yield_mode": div_yield_mode,
            "decomp": decomp,
            "yield_input": yield_input,
            "adv_3m": num(ws[f"EN{r}"].value),
            "perf": {
                "m1": num(ws[f"EB{r}"].value),
                "m3": num(ws[f"EC{r}"].value),
                "m6": num(ws[f"ED{r}"].value),
            },
            "model": model,
        }
        assert current and current != "INDEX"
        current["companies"].append(company)

    payload = {
        "source": "Detailed Dashboard Summary tab",
        "as_of": dt.date.today().isoformat(),
        "base_year": 2026,
        "groups": groups,
        "indexes": indexes,
    }
    with open(dest, "w") as f:
        json.dump(payload, f, indent=1)
    n = sum(len(g["companies"]) for g in groups)
    print(f"Wrote {dest}: {n} companies in {len(groups)} groups, {len(indexes)} indexes")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
