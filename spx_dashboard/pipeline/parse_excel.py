"""
Parse the `Output` sheet of SPX_inputs.xlsx into a single dashboard.json.

The Output sheet is fully self-contained: when the workbook is saved from
Bloomberg, every formula carries its last cached value, so we open with
data_only=True and never touch a terminal.

Tables are located by *searching for their title strings* rather than by
hard-coded row numbers, so small vertical shifts when the user refreshes the
file do not break parsing. Column positions within each table are stable
(they come from a fixed template) and are addressed by letter.

Usage:
    python parse_excel.py <input.xlsx> <output.json>
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string as cidx


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _num(v: Any) -> float | None:
    """Coerce a cell value to float, or None if it is not numeric."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("%", ""))
    except (TypeError, ValueError):
        return None


def _text(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _fmt_date(v: Any) -> str | None:
    """Render a cell date as e.g. '1/1/26'."""
    if isinstance(v, (dt.datetime, dt.date)):
        return f"{v.month}/{v.day}/{str(v.year)[-2:]}"
    return _text(v) or None


def _iso_date(v: Any) -> str | None:
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    return None


def _find_title(ws, needle: str, exact: bool = False) -> tuple[int, int] | None:
    """Return (row, col) of the first cell matching `needle` (substring, or
    exact case-insensitive match when exact=True)."""
    needle = needle.lower()
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            text = str(cell.value).strip().lower()
            if (text == needle) if exact else (needle in text):
                return cell.row, cell.column
    return None


def _cell(ws, row: int, col: "str | int") -> Any:
    col_idx = col if isinstance(col, int) else cidx(col)
    return ws.cell(row=row, column=col_idx).value


# --------------------------------------------------------------------------- #
# Financial tables that share the "3 dates + $Δ + %Δ" shape
# --------------------------------------------------------------------------- #
# Stock Performance, 2026 Estimates and 2027 Estimates all use this layout:
#   label in col B; values D/E/F; $Δ H/I; %Δ K/L
THREE_DATE_VALUE_COLS = ["D", "E", "F"]
THREE_DATE_DELTA_ABS = ["H", "I"]
THREE_DATE_DELTA_PCT = ["K", "L"]


def parse_three_date_table(ws, title_needle: str, value_label: str) -> dict:
    pos = _find_title(ws, title_needle)
    if pos is None:
        raise ValueError(f"Could not find table title containing {title_needle!r}")
    title_row, label_col_idx = pos
    label_col = openpyxl.utils.get_column_letter(label_col_idx)
    title = _text(ws.cell(row=title_row, column=label_col_idx).value)

    # Header rows: title+2 holds the group labels, title+3 holds the dates / Δ labels.
    date_row = title_row + 3
    dates = [_fmt_date(_cell(ws, date_row, c)) for c in THREE_DATE_VALUE_COLS]
    dates_iso = [_iso_date(_cell(ws, date_row, c)) for c in THREE_DATE_VALUE_COLS]

    main_rows: list[dict] = []
    pct_rows: list[dict] = []

    # Scan a generous window below the header.
    for r in range(title_row + 5, title_row + 80):
        label = _text(_cell(ws, r, label_col))
        if not label:
            continue
        is_pct_section = "as % of spx" in label.lower()
        values = [_num(_cell(ws, r, c)) for c in THREE_DATE_VALUE_COLS]
        if all(v is None for v in values):
            # blank gap inside the table; stop if we've left both sections
            if main_rows and is_pct_section is False and pct_rows:
                break
            continue

        row_obj = {
            "label": label.replace(" as % of SPX", ""),
            "values": values,
            "delta_abs": [_num(_cell(ws, r, c)) for c in THREE_DATE_DELTA_ABS],
            "delta_pct": [_num(_cell(ws, r, c)) for c in THREE_DATE_DELTA_PCT],
            "is_total": label.lower().startswith("total"),
        }
        if is_pct_section:
            # pct-of-spx companion: only D/E/F + H/I (share + change) are meaningful
            pct_rows.append(row_obj)
        else:
            main_rows.append(row_obj)

        if "bloomberg spx" in label.lower():
            break

    return {
        "title": title,
        "value_label": value_label,
        "dates": dates,
        "dates_iso": dates_iso,
        "rows": main_rows,
        "pct_of_spx": pct_rows,
    }


# --------------------------------------------------------------------------- #
# Net Income Growth (Earnings Growth) — 2024..2027 + $Δ yoy + %Δ yoy
# --------------------------------------------------------------------------- #
def parse_growth_table(ws) -> dict:
    pos = _find_title(ws, "Net Income Growth")
    if pos is None:
        raise ValueError("Could not find 'Net Income Growth' table")
    title_row, label_col_idx = pos
    label_col = openpyxl.utils.get_column_letter(label_col_idx)
    title = _text(ws.cell(row=title_row, column=label_col_idx).value)

    hdr_row = title_row + 3  # years
    years = [_text(_cell(ws, hdr_row, c)) for c in ("N", "O", "P", "Q")]
    delta_years = [_text(_cell(ws, hdr_row, c)) for c in ("S", "T", "U")]

    main_rows: list[dict] = []
    pct_rows: list[dict] = []
    for r in range(title_row + 5, title_row + 80):
        label = _text(_cell(ws, r, label_col))
        if not label:
            continue
        is_pct = "as % of spx" in label.lower()
        values = [_num(_cell(ws, r, c)) for c in ("N", "O", "P", "Q")]
        if all(v is None for v in values):
            if main_rows and pct_rows:
                break
            continue
        row_obj = {
            "label": label.replace(" as % of SPX", ""),
            "values": values,
            "delta_abs": [_num(_cell(ws, r, c)) for c in ("S", "T", "U")],
            "delta_pct": [_num(_cell(ws, r, c)) for c in ("W", "X", "Y")],
            "is_total": label.lower().startswith("total"),
        }
        (pct_rows if is_pct else main_rows).append(row_obj)

    return {
        "title": title,
        "value_label": "Adj. Net Income",
        "years": years,
        "delta_years": delta_years,
        "rows": main_rows,
        "pct_of_spx": pct_rows,
    }


# --------------------------------------------------------------------------- #
# NTM P/E
# --------------------------------------------------------------------------- #
def parse_ntm_pe(ws) -> dict:
    pos = _find_title(ws, "NTM P/E")
    if pos is None:
        raise ValueError("Could not find 'NTM P/E' table")
    title_row, label_col_idx = pos
    label_col = openpyxl.utils.get_column_letter(label_col_idx)

    # header row with the 'avg since' / time-series dates
    hdr = title_row + 3  # row 189 when title is 186
    avg_cols = ["AV", "AW", "AX", "AY"]
    delta_cols = ["BA", "BB", "BC", "BD"]
    avg_dates = [_fmt_date(_cell(ws, hdr, c)) for c in avg_cols]
    avg_dates_iso = [_iso_date(_cell(ws, hdr, c)) for c in avg_cols]

    # historical quarterly series spans BF.. to the last non-empty header cell
    series_cols: list[str] = []
    series_dates: list[str | None] = []
    c = cidx("BF")
    while True:
        letter = openpyxl.utils.get_column_letter(c)
        d = _cell(ws, hdr, c)
        if d is None or _fmt_date(d) is None:
            break
        series_cols.append(letter)
        series_dates.append(_iso_date(d))
        c += 1

    current_label = _text(_cell(ws, title_row + 2, "AP")) or "Current"

    rows: list[dict] = []
    for r in range(title_row + 5, title_row + 40):
        label = _text(_cell(ws, r, label_col))
        if not label:
            continue
        rows.append(
            {
                "label": label,
                "mkt_cap": _num(_cell(ws, r, "AP")),
                "ntm_ni": _num(_cell(ws, r, "AR")),
                "ntm_pe": _num(_cell(ws, r, "AT")),
                "avg_since": [_num(_cell(ws, r, c)) for c in avg_cols],
                "delta_vs_avg": [_num(_cell(ws, r, c)) for c in delta_cols],
                "series": [_num(_cell(ws, r, c)) for c in series_cols],
                "is_total": label.lower().startswith(("total", "bloomberg")),
            }
        )
        if "bloomberg spx" in label.lower():
            break

    return {
        "title": "NTM P/E",
        "current_label": current_label,
        "avg_dates": avg_dates,
        "avg_dates_iso": avg_dates_iso,
        "series_dates": series_dates,
        "rows": rows,
    }


# --------------------------------------------------------------------------- #
# Per-stock universe (the Data sheet) — powers the category drill-down pages
# --------------------------------------------------------------------------- #
# The `Data` sheet carries one row per S&P 500 constituent. The category rows
# on the `Output` sheet are sums of these per-stock rows, so we read the same
# metric blocks here (calendarized adj. net income, matching the Output tables)
# and key them by company name to attach to each category.
def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def _sub(a: float | None, b: float | None) -> float | None:
    return None if (a is None or b is None) else a - b


def _growth(a: float | None, b: float | None) -> float | None:
    """Change from b to a, as a fraction of |b| (YoY growth / revision)."""
    return None if (a is None or b is None or b == 0) else (a - b) / abs(b)


def _scaled(ws, row: int, col: str, div: float = 1.0) -> float | None:
    v = _num(_cell(ws, row, col))
    return None if v is None else v / div


def _series_cols(ws, start_letter: str) -> list[int]:
    """Walk a horizontal date-headed block (row 10 holds the dates, row 9 the
    block title only on the first column) and return its column indices."""
    cols: list[int] = []
    start = cidx(start_letter)
    c = start
    while True:
        d = ws.cell(row=10, column=c).value
        if not isinstance(d, (dt.datetime, dt.date)):
            break
        if c != start and ws.cell(row=9, column=c).value:
            break  # ran into the next block
        cols.append(c)
        c += 1
    return cols


def _norm_ticker(v: Any) -> str:
    return str(v).strip().upper() if v is not None else ""


def read_compounder_tickers(wb) -> set[str]:
    """The authoritative compounder list lives on the `Compounders` sheet
    (column A, Bloomberg tickers). The Data sheet's "Is compounder?" column is
    just a cached =MATCH() against this list and can go stale, so we match
    against the source directly."""
    if "Compounders" not in wb.sheetnames:
        return set()
    ws = wb["Compounders"]
    out: set[str] = set()
    for r in range(1, ws.max_row + 1):
        t = _norm_ticker(ws.cell(row=r, column=1).value)
        if t and t != "TICKER":
            out.add(t)
    return out


def parse_stock_universe(ws, compounders: set[str] | None = None) -> dict[str, dict]:
    """Read the Data sheet into {company name -> per-stock metrics}."""
    compounders = compounders or set()
    ni_cols = _series_cols(ws, "CT")    # NTM net income estimate history ($m)
    mkt_cols = _series_cols(ws, "DR")   # market cap history ($b), same dates
    n = min(len(ni_cols), len(mkt_cols))

    def _est(r: int, c0: str, c1: str, c2: str) -> dict:
        vv = [_scaled(ws, r, c, 1000) for c in (c0, c1, c2)]
        return {
            "values": vv,
            "delta_abs": [_sub(vv[2], vv[0]), _sub(vv[2], vv[1])],
            "delta_pct": [_growth(vv[2], vv[0]), _growth(vv[2], vv[1])],
        }

    universe: dict[str, dict] = {}
    for r in range(11, ws.max_row + 1):
        name = _text(_cell(ws, r, "C"))
        if not name:
            continue

        # Performance: market cap ($b) at 3 dates, $Δ (YTD/QTD), return % (YTD/QTD)
        perf = {
            "values": [_num(_cell(ws, r, c)) for c in ("AA", "AB", "AC")],
            "delta_abs": [_num(_cell(ws, r, c)) for c in ("AE", "AF")],
            "delta_pct": [_num(_cell(ws, r, c)) for c in ("AH", "AI")],
        }

        # Earnings growth: calendarized adj. net income ($b), 2024..2027 + YoY Δ
        ev = [_scaled(ws, r, c, 1000) for c in ("BO", "BP", "BT", "BX")]
        earnings = {
            "values": ev,
            "delta_abs": [_sub(ev[i + 1], ev[i]) for i in range(3)],
            "delta_pct": [_growth(ev[i + 1], ev[i]) for i in range(3)],
        }

        # NTM P/E: current + quarterly history (market cap / NTM net income)
        ni_hist = [_num(ws.cell(row=r, column=c).value) for c in ni_cols[:n]]
        mkt_hist = [_num(ws.cell(row=r, column=c).value) for c in mkt_cols[:n]]
        pe_series = [
            (mkt_hist[i] * 1000 / ni_hist[i])
            if (mkt_hist[i] is not None and ni_hist[i])
            else None
            for i in range(n)
        ]
        ntm_ni = ni_hist[0] if ni_hist else None

        ticker = _text(_cell(ws, r, "B"))
        universe[name] = {
            "name": name,
            "ticker": ticker,
            "is_compounder": _norm_ticker(ticker) in compounders,
            "performance": perf,
            "earnings": earnings,
            "est_2026": _est(r, "BR", "BS", "BT"),
            "est_2027": _est(r, "BV", "BW", "BX"),
            "pe": {
                "mkt_cap": perf["values"][2],
                "ntm_ni": None if ntm_ni is None else ntm_ni / 1000,
                "ntm_pe": pe_series[0] if pe_series else None,
                "series": pe_series,
            },
        }
    return universe


# --------------------------------------------------------------------------- #
# SPX Categories — the universe map
# --------------------------------------------------------------------------- #
KNOWN_CATEGORIES = {
    "digital semis",
    "analog/mcu",
    "semicap",
    "hardware/components",
    "electric/cooling",
    "power",
    "memory",
    "design",
    "construction",
    "application",
    "infrastructure",
    "big tech",
    "miscellaneous",
}

GROUP_HEADERS = {
    "ai capex beneficiaries",
    "software",
    "ai buildout funders",
    "other",
}

# group header -> the name columns that feed it
GROUP_NAME_COLS = {
    "AI Capex Beneficiaries": ["AB", "AD", "AF"],
    "Software": ["AJ"],
    "AI Buildout Funders": ["AJ"],
    "Other": ["AL"],
}


def parse_categories(ws, universe: dict | None = None) -> dict:
    pos = _find_title(ws, "AI Capex Beneficiaries", exact=True)
    if pos is None:
        raise ValueError("Could not find 'AI Capex Beneficiaries' map")
    top_row = pos[0]  # row 154

    # Walk every name-bearing column, grouping members under the most recent
    # category header seen in that column.
    name_cols = ["AB", "AD", "AF", "AJ", "AL"]
    categories: dict[str, list[str]] = {}
    order: list[str] = []

    for col in name_cols:
        current: str | None = None
        blanks = 0
        for r in range(top_row, top_row + 40):
            val = _text(_cell(ws, r, col))
            if not val:
                blanks += 1
                if blanks > 4 and current:
                    pass
                continue
            blanks = 0
            if val.lower() in GROUP_HEADERS:
                continue  # group-level header, not a category or a member
            if val.lower() in KNOWN_CATEGORIES:
                current = val
                if current not in categories:
                    categories[current] = []
                    order.append(current)
            elif current is not None:
                categories[current].append(val)

    # Map each category to its parent group for display grouping.
    parent = {
        "Digital Semis": "AI Capex Beneficiaries",
        "Memory": "AI Capex Beneficiaries",
        "Semicap": "AI Capex Beneficiaries",
        "Analog/MCU": "AI Capex Beneficiaries",
        "Power": "AI Capex Beneficiaries",
        "Electric/Cooling": "AI Capex Beneficiaries",
        "Hardware/Components": "AI Capex Beneficiaries",
        "Design": "AI Capex Beneficiaries",
        "Construction": "AI Capex Beneficiaries",
        "Application": "Software",
        "Infrastructure": "Software",
        "Big Tech": "AI Buildout Funders",
        "Miscellaneous": "Other",
    }

    # Names explicitly placed in a named category. Everything else in the Data
    # sheet universe falls into "Miscellaneous" (the rest of the S&P 500).
    assigned = {
        m
        for cat in order
        if cat.lower() != "miscellaneous"
        for m in categories[cat]
        if universe and m in universe
    }

    groups: dict[str, list[dict]] = {}
    for cat in order:
        g = parent.get(cat, "Other")
        if cat.lower() == "miscellaneous" and universe:
            # All remaining constituents, kept in Data-sheet order (largest first).
            members = [name for name in universe if name not in assigned]
        else:
            members = categories[cat]
        stocks = [universe[m] for m in members if universe and m in universe]
        groups.setdefault(g, []).append(
            {
                "category": cat,
                "slug": _slug(cat),
                "members": members,
                "stocks": stocks,
            }
        )

    group_order = ["AI Capex Beneficiaries", "Software", "AI Buildout Funders", "Other"]
    return {
        "title": "SPX Categories",
        "groups": [
            {"group": g, "categories": groups[g]} for g in group_order if g in groups
        ],
    }


# --------------------------------------------------------------------------- #
# Subset roll-ups (e.g. "compounders only")
# --------------------------------------------------------------------------- #
# Every aggregate row on the Output sheet is a SUMIFS over the Data sheet,
# filtered by a flag column: P = AI-capex sub-category, G/F/H/I/J = the broad
# buckets, all rows = Total SPX. Restricting to compounders simply adds the
# Data!D="yes" criterion to each sum. We replicate that here so any subset can
# be rolled up with the exact same arithmetic the workbook uses.
_LABEL_SELECTOR: dict[str, tuple[str, str | None]] = {
    "Total AI Capex Beneficiaries": ("FLAG", "G"),
    "AI Buildout Funders": ("FLAG", "F"),
    "Infrastructure software": ("FLAG", "H"),
    "Application software": ("FLAG", "I"),
    "Other": ("FLAG", "J"),
    "Total SPX": ("ALL", None),
}


def _selector_for(label: str) -> tuple[str, str | None] | None:
    if label in _LABEL_SELECTOR:
        return _LABEL_SELECTOR[label]
    if label.lower().startswith("bloomberg"):
        return None  # external benchmark; not derivable from constituents
    return ("P", label)  # an AI-capex sub-category, matched on Data!P


def read_agg_records(ws, compounders: set[str] | None = None) -> dict:
    """Read the Data sheet into per-stock roll-up records + the P/E history dates."""
    compounders = compounders or set()
    ni_cols = _series_cols(ws, "CT")   # NTM net income history ($m)
    mkt_cols = _series_cols(ws, "DR")  # market cap history ($b)
    n = min(len(ni_cols), len(mkt_cols))
    hist_dates = [_iso_date(ws.cell(row=10, column=c).value) for c in ni_cols[:n]]

    records: list[dict] = []
    for r in range(11, ws.max_row + 1):
        name = _text(_cell(ws, r, "C"))
        if not name:
            continue
        flag = lambda col: _text(_cell(ws, r, col)).lower() == "yes"  # noqa: E731
        records.append(
            {
                "name": name,
                "comp": _norm_ticker(_cell(ws, r, "B")) in compounders,
                "P": _text(_cell(ws, r, "P")),
                "F": flag("F"),
                "G": flag("G"),
                "H": flag("H"),
                "I": flag("I"),
                "J": flag("J"),
                "perf": [_num(_cell(ws, r, c)) for c in ("AA", "AB", "AC")],
                "earn": [_num(_cell(ws, r, c)) for c in ("BO", "BP", "BT", "BX")],
                "e26": [_num(_cell(ws, r, c)) for c in ("BR", "BS", "BT")],
                "e27": [_num(_cell(ws, r, c)) for c in ("BV", "BW", "BX")],
                "ni_hist": [_num(ws.cell(row=r, column=c).value) for c in ni_cols[:n]],
                "mkt_hist": [_num(ws.cell(row=r, column=c).value) for c in mkt_cols[:n]],
            }
        )
    return {"records": records, "hist_dates": hist_dates}


def _members(records: list[dict], label: str, compounders: bool) -> list[dict] | None:
    sel = _selector_for(label)
    if sel is None:
        return None
    kind, arg = sel
    out = []
    for rec in records:
        if compounders and not rec["comp"]:
            continue
        if kind == "P":
            ok = rec["P"] == label
        elif kind == "FLAG":
            ok = rec[arg]
        else:  # ALL
            ok = True
        if ok:
            out.append(rec)
    return out


def _sum(vals) -> float | None:
    xs = [v for v in vals if v is not None]
    return sum(xs) if xs else None


def _div(a: float | None, b: float | None) -> float | None:
    return None if (a is None or b in (None, 0)) else a / b


def _pe_val(mkt: float | None, ni_m: float | None) -> float | None:
    """P/E from market cap ($b) and net income ($m); meaningless when earnings
    aren't positive, so those periods are dropped rather than shown as garbage."""
    if mkt is None or ni_m is None or ni_m <= 0:
        return None
    return mkt * 1000 / ni_m


def _residual(total: float | None, parts: list) -> float | None:
    """Total minus the named buckets — how the Output sheet defines 'Other'.
    A bucket with no members contributes 0 (not None) to the subtraction."""
    if total is None:
        return None
    return total - sum(p for p in parts if p is not None)


# Named buckets whose sum, subtracted from Total SPX, yields "Other".
_OTHER_PARTS = [
    "Total AI Capex Beneficiaries",
    "AI Buildout Funders",
    "Infrastructure software",
    "Application software",
]


def _three_date_subset(base: dict, records, compounders: bool, key: str, div: float) -> dict:
    raw: dict[str, list] = {}
    for br in base["rows"]:
        mem = _members(records, br["label"], compounders)
        if mem is None:
            continue
        raw[br["label"]] = [
            (lambda s: None if s is None else s / div)(_sum([rec[key][i] for rec in mem]))
            for i in range(3)
        ]
    if "Other" in raw and "Total SPX" in raw:
        raw["Other"] = [
            _residual(raw["Total SPX"][i], [raw[n][i] for n in _OTHER_PARTS if n in raw])
            for i in range(3)
        ]

    rows = []
    for br in base["rows"]:
        if br["label"] not in raw:
            continue
        vals = raw[br["label"]]
        d0 = None if (vals[2] is None or vals[0] is None) else vals[2] - vals[0]
        d1 = None if (vals[2] is None or vals[1] is None) else vals[2] - vals[1]
        p0 = _div(vals[2], vals[0])
        p1 = _div(vals[2], vals[1])
        rows.append(
            {
                "label": br["label"],
                "values": vals,
                "delta_abs": [d0, d1],
                "delta_pct": [None if p0 is None else p0 - 1, None if p1 is None else p1 - 1],
                "is_total": br["is_total"],
            }
        )
    meta = {k: base[k] for k in ("title", "value_label", "dates", "dates_iso")}
    return {**meta, "rows": rows, "pct_of_spx": []}


def _growth_subset(base: dict, records, compounders: bool) -> dict:
    raw: dict[str, list] = {}
    for br in base["rows"]:
        mem = _members(records, br["label"], compounders)
        if mem is None:
            continue
        raw[br["label"]] = [
            (lambda s: None if s is None else s / 1000)(_sum([rec["earn"][i] for rec in mem]))
            for i in range(4)
        ]
    if "Other" in raw and "Total SPX" in raw:
        raw["Other"] = [
            _residual(raw["Total SPX"][i], [raw[n][i] for n in _OTHER_PARTS if n in raw])
            for i in range(4)
        ]

    rows = []
    for br in base["rows"]:
        if br["label"] not in raw:
            continue
        vals = raw[br["label"]]
        dabs = [
            None if (vals[i + 1] is None or vals[i] is None) else vals[i + 1] - vals[i]
            for i in range(3)
        ]
        dpct = []
        for i in range(3):
            q = _div(vals[i + 1], vals[i])
            dpct.append(None if q is None else q - 1)
        rows.append(
            {
                "label": br["label"],
                "values": vals,
                "delta_abs": dabs,
                "delta_pct": dpct,
                "is_total": br["is_total"],
            }
        )
    meta = {k: base[k] for k in ("title", "value_label", "years", "delta_years")}
    return {**meta, "rows": rows, "pct_of_spx": []}


def _pe_subset(base: dict, agg: dict, compounders: bool) -> dict:
    records = agg["records"]
    hist_dates = agg["hist_dates"]
    series_dates = base["series_dates"]
    avg_iso = base.get("avg_dates_iso") or []
    # AVERAGEIFS uses ">=" & (avg_date - 1 day) on the history date headers.
    thresholds = [
        (dt.date.fromisoformat(x) - dt.timedelta(days=1)) if x else None for x in avg_iso
    ]
    n = len(hist_dates)

    # Raw market-cap and net-income aggregates per row (so "Other" can be taken
    # as a residual at the numerator/denominator level, like the workbook does).
    raw: dict[str, dict] = {}
    for br in base["rows"]:
        mem = _members(records, br["label"], compounders)
        if mem is None:
            continue  # external benchmark rows (Bloomberg SPX) are dropped
        raw[br["label"]] = {
            "mkt_cap": _sum([rec["perf"][2] for rec in mem]),     # current mkt cap (AC)
            "ni_cur": _sum([rec["ni_hist"][0] for rec in mem]),   # current NTM NI ($m)
            "mkt_h": [_sum([rec["mkt_hist"][k] for rec in mem]) for k in range(n)],
            "ni_h": [_sum([rec["ni_hist"][k] for rec in mem]) for k in range(n)],
        }
    if "Other" in raw and "Total SPX" in raw:
        T = raw["Total SPX"]
        parts = [raw[p] for p in _OTHER_PARTS if p in raw]
        o = raw["Other"]
        o["mkt_cap"] = _residual(T["mkt_cap"], [p["mkt_cap"] for p in parts])
        o["ni_cur"] = _residual(T["ni_cur"], [p["ni_cur"] for p in parts])
        o["mkt_h"] = [_residual(T["mkt_h"][k], [p["mkt_h"][k] for p in parts]) for k in range(n)]
        o["ni_h"] = [_residual(T["ni_h"][k], [p["ni_h"][k] for p in parts]) for k in range(n)]

    rows = []
    for br in base["rows"]:
        if br["label"] not in raw:
            continue
        R = raw[br["label"]]
        ntm_ni = None if R["ni_cur"] is None else R["ni_cur"] / 1000
        ntm_pe = _pe_val(R["mkt_cap"], R["ni_cur"])

        pe_by_date = {hist_dates[k]: _pe_val(R["mkt_h"][k], R["ni_h"][k]) for k in range(n)}
        series = [pe_by_date.get(d) for d in series_dates]
        avg_since = []
        for thr in thresholds:
            if thr is None:
                avg_since.append(None)
                continue
            vals = [
                v
                for d, v in pe_by_date.items()
                if v is not None and d and dt.date.fromisoformat(d) >= thr
            ]
            avg_since.append(sum(vals) / len(vals) if vals else None)
        delta = []
        for a in avg_since:
            q = _div(ntm_pe, a)
            delta.append(None if q is None else q - 1)

        rows.append(
            {
                "label": br["label"],
                "mkt_cap": R["mkt_cap"],
                "ntm_ni": ntm_ni,
                "ntm_pe": ntm_pe,
                "avg_since": avg_since,
                "delta_vs_avg": delta,
                "series": series,
                "is_total": br["is_total"],
            }
        )
    meta = {
        k: base[k]
        for k in ("title", "current_label", "avg_dates", "avg_dates_iso", "series_dates")
    }
    return {**meta, "rows": rows}


def build_subset_tables(agg: dict, base_tables: dict, compounders: bool) -> dict:
    recs = agg["records"]
    return {
        "stock_performance": _three_date_subset(
            base_tables["stock_performance"], recs, compounders, "perf", 1
        ),
        "est_rev_2026": _three_date_subset(
            base_tables["est_rev_2026"], recs, compounders, "e26", 1000
        ),
        "est_rev_2027": _three_date_subset(
            base_tables["est_rev_2027"], recs, compounders, "e27", 1000
        ),
        "earnings_growth": _growth_subset(base_tables["earnings_growth"], recs, compounders),
        "ntm_pe": _pe_subset(base_tables["ntm_pe"], agg, compounders),
    }


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def read_bloomberg_date(wb) -> str | None:
    """Date of the latest Bloomberg data, read from cell Data!AC10.

    This is the single source of truth for the "data as of" date shown across
    the dashboard. Returns an ISO yyyy-mm-dd string, or None if unavailable.
    """
    if "Data" not in wb.sheetnames:
        return None
    val = wb["Data"]["AC10"].value
    if isinstance(val, dt.datetime):
        return val.date().isoformat()
    if isinstance(val, dt.date):
        return val.isoformat()
    return None


def parse_workbook(path: str, refreshed_date: str | None = None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Output" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Output' sheet")
    ws = wb["Output"]
    compounders = read_compounder_tickers(wb)
    universe = (
        parse_stock_universe(wb["Data"], compounders)
        if "Data" in wb.sheetnames
        else {}
    )

    # When refreshed_date is provided, relabel the last "current" date column
    # in each table to reflect when the file was actually refreshed, since the
    # Excel cell often lags by a few days.
    refreshed_label: str | None = None
    if refreshed_date:
        try:
            d = dt.date.fromisoformat(refreshed_date)
            refreshed_label = f"{d.month}/{d.day}/{str(d.year)[-2:]}"
        except ValueError:
            pass

    def _relabel(table: dict) -> dict:
        if refreshed_label and table.get("dates"):
            table["dates"][-1] = refreshed_label
        return table

    stock_perf = _relabel(parse_three_date_table(
        ws, "YTD Stock Performance", "Market cap ($b)"
    ))

    tables = {
        "stock_performance": stock_perf,
        "est_rev_2026": _relabel(parse_three_date_table(
            ws, "2026 Estimates", "Consensus Adj. Net Income ($b)"
        )),
        "est_rev_2027": _relabel(parse_three_date_table(
            ws, "2027 Estimates", "Consensus Adj. Net Income ($b)"
        )),
        "earnings_growth": parse_growth_table(ws),
        "ntm_pe": parse_ntm_pe(ws),
        "categories": parse_categories(ws, universe),
    }

    # Compounders-only roll-ups of every aggregate table, computed with the
    # same SUMIFS logic the Output sheet uses (Data!D="yes" added to each sum).
    # The base tables are already relabeled above, so the date headers the
    # subset builder copies are correct.
    tables_compounders: dict = {}
    if "Data" in wb.sheetnames:
        agg = read_agg_records(wb["Data"], compounders)
        tables_compounders = build_subset_tables(agg, tables, compounders=True)

    return {
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(),
        "refreshed_date": refreshed_date,
        # Date of the latest Bloomberg data (Data!AC10) — the value quoted as
        # "data as of" across the dashboard.
        "bloomberg_date": read_bloomberg_date(wb),
        "latest_date": stock_perf["dates"][-1],
        "tables": tables,
        "tables_compounders": tables_compounders,
    }


def main() -> int:
    if len(sys.argv) not in (3, 4):
        print(
            "usage: parse_excel.py <input.xlsx> <output.json> [refreshed_date]",
            file=sys.stderr,
        )
        return 2
    refreshed_date = sys.argv[3] if len(sys.argv) == 4 else None
    data = parse_workbook(sys.argv[1], refreshed_date)
    with open(sys.argv[2], "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"Wrote {sys.argv[2]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
